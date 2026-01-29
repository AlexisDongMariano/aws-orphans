"""
FastAPI app: orphaned security groups from Postgres.
Data is populated by running: python scripts/populate_orphaned_sgs_db.py
Templates live in templates/ (project root).
"""

from io import BytesIO
from pathlib import Path

from fastapi import FastAPI, Request
from fastapi.responses import HTMLResponse, Response
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook

from app.regions import REGIONS
from app.db import get_connection, fetch_all, fetch_all_eips, fetch_all_ebs


# Templates directory: project root / templates (next to app/)
BASE_DIR = Path(__file__).resolve().parent.parent
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))

app = FastAPI(title="AWS Orphans", version="0.1.0")


def get_orphaned_sgs_from_db() -> list[dict]:
    with get_connection() as conn:
        return fetch_all(conn)


def get_orphaned_eips_from_db() -> list[dict]:
    with get_connection() as conn:
        return fetch_all_eips(conn)


def get_orphaned_ebs_from_db() -> list[dict]:
    with get_connection() as conn:
        return fetch_all_ebs(conn)


@app.get("/")
def root():
    return {"service": "aws-orphans", "orphaned_sgs": "/orphaned-sgs", "orphaned_eips": "/orphaned-eips", "orphaned_ebs": "/orphaned-ebs"}


@app.get("/api/regions")
def list_regions():
    return {"regions": REGIONS}


@app.get("/api/orphaned-sgs")
def api_orphaned_sgs():
    """JSON list of orphaned SGs from the database."""
    return get_orphaned_sgs_from_db()


@app.get("/api/orphaned-sgs/export")
def api_orphaned_sgs_export():
    """Download the table as an Excel file."""
    rows = get_orphaned_sgs_from_db()
    wb = Workbook()
    ws = wb.active
    ws.title = "Orphaned SGs"
    headers = ["Region", "Group ID", "Name", "Description", "VPC", "Console URL"]
    ws.append(headers)
    for r in rows:
        ws.append([
            r.get("region", ""),
            r.get("group_id", ""),
            r.get("group_name", ""),
            r.get("description", ""),
            r.get("vpc_id") or "",
            r.get("console_url", ""),
        ])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=orphaned-security-groups.xlsx"},
    )


@app.get("/api/orphaned-sgs/table", response_class=HTMLResponse)
def api_orphaned_sgs_table_fragment(request: Request):
    """HTML fragment: stats + full table. Used by HTMX on /orphaned-sgs."""
    rows = get_orphaned_sgs_from_db()
    total = len(rows)
    by_region: dict[str, int] = {}
    for r in rows:
        by_region[r["region"]] = by_region.get(r["region"], 0) + 1
    regions_with_count = [{"region": reg, "count": n} for reg, n in sorted(by_region.items())]
    return templates.TemplateResponse(
        "orphaned_sgs_table.html",
        {
            "request": request,
            "rows": rows,
            "total": total,
            "regions_with_count": regions_with_count,
        },
    )


@app.get("/orphaned-sgs", response_class=HTMLResponse)
def page_orphaned_sgs(request: Request):
    """HTML page with Tailwind + HTMX: table of orphaned SGs loaded from DB."""
    return templates.TemplateResponse("orphaned_sgs.html", {"request": request, "active": "sgs"})


# --- Orphaned Elastic IPs ---

@app.get("/api/orphaned-eips")
def api_orphaned_eips():
    """JSON list of orphaned EIPs from the database."""
    return get_orphaned_eips_from_db()


@app.get("/api/orphaned-eips/export")
def api_orphaned_eips_export():
    """Download the EIP table as an Excel file."""
    rows = get_orphaned_eips_from_db()
    wb = Workbook()
    ws = wb.active
    ws.title = "Orphaned EIPs"
    headers = ["Region", "Allocation ID", "Public IP", "Domain", "Console URL"]
    ws.append(headers)
    for r in rows:
        ws.append([
            r.get("region", ""),
            r.get("allocation_id", ""),
            r.get("public_ip", ""),
            r.get("domain", ""),
            r.get("console_url", ""),
        ])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=orphaned-elastic-ips.xlsx"},
    )


@app.get("/api/orphaned-eips/table", response_class=HTMLResponse)
def api_orphaned_eips_table_fragment(request: Request):
    """HTML fragment: stats + full table. Used by HTMX on /orphaned-eips."""
    rows = get_orphaned_eips_from_db()
    total = len(rows)
    by_region: dict[str, int] = {}
    for r in rows:
        by_region[r["region"]] = by_region.get(r["region"], 0) + 1
    regions_with_count = [{"region": reg, "count": n} for reg, n in sorted(by_region.items())]
    return templates.TemplateResponse(
        "orphaned_eips_table.html",
        {
            "request": request,
            "rows": rows,
            "total": total,
            "regions_with_count": regions_with_count,
        },
    )


@app.get("/orphaned-eips", response_class=HTMLResponse)
def page_orphaned_eips(request: Request):
    """HTML page with Tailwind + HTMX: table of orphaned EIPs loaded from DB."""
    return templates.TemplateResponse("orphaned_eips.html", {"request": request, "active": "eips"})


# --- Unattached EBS Volumes ---

@app.get("/api/orphaned-ebs")
def api_orphaned_ebs():
    """JSON list of unattached EBS volumes from the database."""
    return get_orphaned_ebs_from_db()


@app.get("/api/orphaned-ebs/export")
def api_orphaned_ebs_export():
    """Download the EBS table as an Excel file."""
    rows = get_orphaned_ebs_from_db()
    wb = Workbook()
    ws = wb.active
    ws.title = "Unattached EBS"
    headers = ["Region", "Volume ID", "Size (GB)", "Type", "Availability Zone", "Created", "Console URL"]
    ws.append(headers)
    for r in rows:
        ws.append([
            r.get("region", ""),
            r.get("volume_id", ""),
            r.get("size_gb", 0),
            r.get("volume_type", ""),
            r.get("availability_zone", ""),
            r.get("create_time", ""),
            r.get("console_url", ""),
        ])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=unattached-ebs-volumes.xlsx"},
    )


@app.get("/api/orphaned-ebs/table", response_class=HTMLResponse)
def api_orphaned_ebs_table_fragment(request: Request):
    """HTML fragment: stats + full table. Used by HTMX on /orphaned-ebs."""
    rows = get_orphaned_ebs_from_db()
    total = len(rows)
    by_region: dict[str, int] = {}
    for r in rows:
        by_region[r["region"]] = by_region.get(r["region"], 0) + 1
    regions_with_count = [{"region": reg, "count": n} for reg, n in sorted(by_region.items())]
    return templates.TemplateResponse(
        "orphaned_ebs_table.html",
        {
            "request": request,
            "rows": rows,
            "total": total,
            "regions_with_count": regions_with_count,
        },
    )


@app.get("/orphaned-ebs", response_class=HTMLResponse)
def page_orphaned_ebs(request: Request):
    """HTML page with Tailwind + HTMX: table of unattached EBS volumes loaded from DB."""
    return templates.TemplateResponse("orphaned_ebs.html", {"request": request, "active": "ebs"})
